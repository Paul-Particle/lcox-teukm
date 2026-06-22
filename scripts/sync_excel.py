"""
sync_excel.py — bidirectional sync between config.yaml + cases.csv and config.xlsx.

The Excel workbook has five sheets that mirror the source files:
  shared      — flat key/value table for the shared economic params
  platforms   — one column per platform,  one row per field path (dot notation)
  drivetrains — one column per drivetrain, one row per field path
  sources     — one column per source,     one row per field path
  cases       — the cases.csv table verbatim

Usage:
    uv run scripts/sync_excel.py [--check | --to-excel | --to-yaml]

    --check      compare Excel with YAML + CSV and report diffs (default when
                 called with no flags; interactive if diffs are found)
    --to-excel   write Excel from YAML + CSV (backs up any existing .xlsx first)
    --to-yaml    write YAML + CSV from Excel (backs up both text files first;
                 preserves all comments in config.yaml via ruamel.yaml)
"""
from __future__ import annotations

import argparse
import math
import shutil
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from ruamel.yaml import YAML

ROOT = Path(__file__).resolve().parent.parent
YAML_PATH  = ROOT / "config.yaml"
CSV_PATH   = ROOT / "cases.csv"
EXCEL_PATH = ROOT / "config.xlsx"
BACKUP_DIR = ROOT / "backups"

COMPONENT_SECTIONS = ("platforms", "drivetrains", "sources")
INT_COLS = ("optimize_n", "sweep_n")   # cases columns that must stay integer


# ── YAML helpers ──────────────────────────────────────────────────────────────

def _yaml() -> YAML:
    y = YAML()
    y.preserve_quotes = True  # keeps single-quoted strings as single-quoted
    return y


def load_yaml(path: Path):
    with open(path) as f:
        return _yaml().load(f)


def save_yaml(data, path: Path) -> None:
    with open(path, "w") as f:
        _yaml().dump(data, f)


# ── dict utilities ────────────────────────────────────────────────────────────

def flatten(d: dict, prefix: str = "") -> dict:
    """Recursively flatten a (possibly ruamel.yaml CommentedMap) to {dot.path: value}."""
    out = {}
    for k, v in d.items():
        key = f"{prefix}.{k}" if prefix else str(k)
        if isinstance(v, dict):
            out.update(flatten(v, key))
        else:
            out[key] = v
    return out


def get_nested(d, path: str) -> tuple[bool, object]:
    """Navigate a nested dict via a dot-path. Returns (found, value)."""
    parts = path.split(".")
    for part in parts[:-1]:
        if not isinstance(d, dict) or part not in d:
            return False, None
        d = d[part]
    leaf = parts[-1]
    return (leaf in d, d.get(leaf))


def set_nested(d, path: str, value) -> None:
    """Set a value at a dot-path in a (possibly ruamel.yaml) nested dict.
    Coerces to float when the existing value is float and the new one is int,
    so ruamel.yaml writes 25.0 not 25 where the original had 25.0."""
    parts = path.split(".")
    for part in parts[:-1]:
        d = d[part]
    leaf = parts[-1]
    existing = d.get(leaf)
    if isinstance(existing, float) and isinstance(value, int):
        value = float(value)
    d[leaf] = value


# ── Excel formatting helpers ──────────────────────────────────────────────────

_HEADER_FONT = Font(bold=True)
_HEADER_FILL = PatternFill("solid", fgColor="DCE6F1")


def _write_header(ws, row: list) -> None:
    ws.append(row)
    for cell in ws[ws.max_row]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL


def _autowidth(ws, min_w: int = 12, max_w: int = 45) -> None:
    for col in ws.columns:
        width = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = max(min_w, min(width + 2, max_w))


# ── YAML + CSV → Excel ────────────────────────────────────────────────────────

def _write_component_sheet(ws, components: dict[str, dict]) -> None:
    """One column per component, union of all field paths as rows."""
    seen: set[str] = set()
    all_fields: list[str] = []
    for flat in components.values():
        for k in flat:
            if k not in seen:
                all_fields.append(k)
                seen.add(k)
    _write_header(ws, ["field"] + list(components.keys()))
    for field in all_fields:
        ws.append([field] + [components[name].get(field) for name in components])


def yaml_csv_to_excel(yaml_path: Path, csv_path: Path, excel_path: Path) -> None:
    data = load_yaml(yaml_path)
    wb = Workbook()
    wb.remove(wb.active)

    # shared — flat key / value
    ws_shared = wb.create_sheet("shared")
    _write_header(ws_shared, ["key", "value"])
    for k, v in flatten(data["shared"]).items():
        ws_shared.append([k, v])

    # platforms / drivetrains / sources — components as columns
    for section in COMPONENT_SECTIONS:
        components = {name: flatten(block) for name, block in data[section].items()}
        _write_component_sheet(wb.create_sheet(section), components)

    # cases — verbatim from CSV
    cases_df = pd.read_csv(csv_path)
    ws_cases = wb.create_sheet("cases")
    _write_header(ws_cases, list(cases_df.columns))
    for _, row in cases_df.iterrows():
        ws_cases.append([None if pd.isna(v) else v for v in row])

    for ws in wb.worksheets:
        ws.freeze_panes = "B2"
        _autowidth(ws)

    wb.save(excel_path)
    print(f"  wrote {excel_path.relative_to(ROOT)}")


# ── Excel → YAML + CSV ────────────────────────────────────────────────────────

def _read_component_sheet(ws) -> dict[str, dict]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}
    names = [h for h in rows[0][1:] if h is not None]
    result: dict[str, dict] = {n: {} for n in names}
    for row in rows[1:]:
        field = row[0]
        if field is None:
            continue
        for name, value in zip(names, row[1:]):
            if value is not None:
                result[name][str(field)] = value
    return result


def excel_to_dicts(excel_path: Path) -> dict:
    wb = load_workbook(excel_path, data_only=True)

    rows = list(wb["shared"].iter_rows(values_only=True))
    shared_flat = {str(r[0]): r[1] for r in rows[1:] if r[0] is not None}

    component_data = {s: _read_component_sheet(wb[s]) for s in COMPONENT_SECTIONS}

    ws_cases = wb["cases"]
    case_rows = list(ws_cases.iter_rows(values_only=True))
    headers = [str(h) for h in case_rows[0]]
    cases = [dict(zip(headers, row)) for row in case_rows[1:]]

    return {"shared": shared_flat, **component_data, "cases": cases}


def excel_to_yaml_csv(excel_path: Path, yaml_path: Path, csv_path: Path) -> None:
    data = load_yaml(yaml_path)
    xl = excel_to_dicts(excel_path)

    # shared
    for path, value in xl["shared"].items():
        found, _ = get_nested(data["shared"], path)
        if found:
            set_nested(data["shared"], path, value)
        else:
            print(f"  WARNING: shared.{path} not found in YAML — skipped")

    # component sections
    for section in COMPONENT_SECTIONS:
        for name, flat in xl[section].items():
            if name not in data[section]:
                print(f"  WARNING: {section}.{name} not in YAML — skipped (add components in config.yaml)")
                continue
            for path, value in flat.items():
                found, _ = get_nested(data[section][name], path)
                if found:
                    set_nested(data[section][name], path, value)
                else:
                    print(f"  WARNING: {section}.{name}.{path} not found in YAML — skipped")

    save_yaml(data, yaml_path)
    print(f"  wrote {yaml_path.relative_to(ROOT)}")

    # cases CSV
    cases_df = pd.DataFrame(xl["cases"])
    for col in INT_COLS:
        if col in cases_df.columns:
            cases_df[col] = pd.to_numeric(cases_df[col], errors="coerce").astype("Int64")
    cases_df.to_csv(csv_path, index=False)
    print(f"  wrote {csv_path.relative_to(ROOT)}")


# ── Diff / check ──────────────────────────────────────────────────────────────

Diff = tuple[str, object, object]  # (label, yaml/csv value, excel value)


def _eq(a, b) -> bool:
    a_null = a is None or (isinstance(a, float) and math.isnan(a))
    b_null = b is None or (isinstance(b, float) and math.isnan(b))
    if a_null and b_null:
        return True
    if a_null or b_null:
        return False
    if isinstance(a, (int, float)) and isinstance(b, (int, float)):
        return math.isclose(float(a), float(b), rel_tol=1e-9, abs_tol=1e-12)
    return str(a).strip() == str(b).strip()


def _diff_flat(prefix: str, yaml_flat: dict, excel_flat: dict) -> list[Diff]:
    diffs = []
    for k in sorted(set(yaml_flat) | set(excel_flat)):
        y, e = yaml_flat.get(k), excel_flat.get(k)
        if not _eq(y, e):
            diffs.append((f"{prefix}.{k}", y, e))
    return diffs


def check_sync(yaml_path: Path, csv_path: Path, excel_path: Path) -> list[Diff]:
    data = load_yaml(yaml_path)
    xl = excel_to_dicts(excel_path)
    diffs: list[Diff] = []

    diffs += _diff_flat("shared", flatten(data["shared"]), xl["shared"])

    for section in COMPONENT_SECTIONS:
        for name in sorted(set(data[section]) | set(xl[section])):
            y_flat = flatten(data[section].get(name, {}))
            e_flat = xl[section].get(name, {})
            diffs += _diff_flat(f"{section}.{name}", y_flat, e_flat)

    cases_df = pd.read_csv(csv_path)
    xl_cases = xl["cases"]
    if len(cases_df) != len(xl_cases):
        diffs.append(("cases.row_count", len(cases_df), len(xl_cases)))
    else:
        for i, (_, csv_row) in enumerate(cases_df.iterrows()):
            xl_row = xl_cases[i]
            for col in cases_df.columns:
                y_val = None if pd.isna(csv_row[col]) else csv_row[col]
                e_val = xl_row.get(col)
                if not _eq(y_val, e_val):
                    diffs.append((f"cases[{i}].{col}", y_val, e_val))

    return diffs


# ── Backup ────────────────────────────────────────────────────────────────────

def backup(path: Path) -> Path:
    BACKUP_DIR.mkdir(exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    dest = BACKUP_DIR / f"{path.stem}_{stamp}{path.suffix}"
    shutil.copy2(path, dest)
    return dest


# ── CLI ───────────────────────────────────────────────────────────────────────

def _print_diffs(diffs: list[Diff]) -> None:
    print(f"\n{len(diffs)} difference(s):")
    for label, y, e in diffs:
        print(f"  {label}")
        print(f"    YAML/CSV : {y!r}")
        print(f"    Excel    : {e!r}")


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__,
                                     formatter_class=argparse.RawDescriptionHelpFormatter)
    grp = parser.add_mutually_exclusive_group()
    grp.add_argument("--to-excel", action="store_true",
                     help="write Excel from YAML + CSV (backs up existing .xlsx)")
    grp.add_argument("--to-yaml",  action="store_true",
                     help="write YAML + CSV from Excel (backs up config.yaml + cases.csv)")
    grp.add_argument("--check",    action="store_true",
                     help="report diffs; offer interactive sync (default)")
    args = parser.parse_args()

    if args.to_excel:
        if EXCEL_PATH.exists():
            dest = backup(EXCEL_PATH)
            print(f"backed up → {dest.name}")
        yaml_csv_to_excel(YAML_PATH, CSV_PATH, EXCEL_PATH)

    elif args.to_yaml:
        if not EXCEL_PATH.exists():
            sys.exit(f"Excel not found: {EXCEL_PATH}")
        for path in (YAML_PATH, CSV_PATH):
            dest = backup(path)
            print(f"backed up → {dest.name}")
        excel_to_yaml_csv(EXCEL_PATH, YAML_PATH, CSV_PATH)

    else:  # --check (default)
        if not EXCEL_PATH.exists():
            print(f"No Excel file at {EXCEL_PATH.name} — run with --to-excel to create it.")
            return
        print("Checking sync…")
        diffs = check_sync(YAML_PATH, CSV_PATH, EXCEL_PATH)
        if not diffs:
            print("All in sync.")
            return
        _print_diffs(diffs)
        try:
            answer = input("\nSync direction? [e = YAML→Excel / y = Excel→YAML / q = quit]: ").strip().lower()
        except (EOFError, KeyboardInterrupt):
            print()
            return
        if answer == "e":
            if EXCEL_PATH.exists():
                dest = backup(EXCEL_PATH)
                print(f"backed up → {dest.name}")
            yaml_csv_to_excel(YAML_PATH, CSV_PATH, EXCEL_PATH)
        elif answer == "y":
            for path in (YAML_PATH, CSV_PATH):
                dest = backup(path)
                print(f"backed up → {dest.name}")
            excel_to_yaml_csv(EXCEL_PATH, YAML_PATH, CSV_PATH)
        else:
            print("No changes made.")


if __name__ == "__main__":
    main()
